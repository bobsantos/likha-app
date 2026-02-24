"""
Report template generation tests (Phase 2, Task 1).

Tests the GET /api/contracts/{id}/report-template endpoint and the
underlying report_template service.

TDD: tests written before implementation.
"""

import io
import os
import pytest
from unittest.mock import AsyncMock, MagicMock, Mock, patch

# Ensure env vars are set before importing anything that triggers app imports
os.environ.setdefault("SUPABASE_URL", "https://test.supabase.co")
os.environ.setdefault("SUPABASE_KEY", "test-anon-key")
os.environ.setdefault("SUPABASE_SERVICE_KEY", "test-service-key")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_db_contract(
    contract_id="contract-123",
    user_id="user-123",
    licensee_name="Sunrise Apparel Co.",
    royalty_rate="8%",
    status="active",
    product_categories=None,
):
    return {
        "id": contract_id,
        "user_id": user_id,
        "status": status,
        "filename": "test.pdf",
        "licensee_name": licensee_name,
        "pdf_url": "https://test.supabase.co/storage/test.pdf",
        "extracted_terms": {},
        "royalty_rate": royalty_rate,
        "royalty_base": "net sales",
        "territories": [],
        "product_categories": product_categories,
        "contract_start_date": "2025-01-01",
        "contract_end_date": "2025-12-31",
        "minimum_guarantee": "0",
        "minimum_guarantee_period": "annually",
        "advance_payment": None,
        "reporting_frequency": "quarterly",
        "storage_path": f"contracts/{user_id}/test.pdf",
        "created_at": "2025-01-01T00:00:00Z",
        "updated_at": "2025-01-01T00:00:00Z",
    }


def _mock_ownership_and_contract(mock_supabase, contract_data):
    """
    Configure the supabase mock so that table("contracts") select chain
    returns contract_data.
    """
    mock_exec = MagicMock()
    mock_exec.execute.return_value = Mock(data=[contract_data])

    mock_eq = MagicMock()
    mock_eq.eq.return_value = mock_exec
    mock_eq.execute.return_value = Mock(data=[contract_data])

    mock_select = MagicMock()
    mock_select.eq.return_value = mock_eq

    mock_table = MagicMock()
    mock_table.select.return_value = mock_select

    mock_supabase.table.return_value = mock_table
    return mock_table


# ---------------------------------------------------------------------------
# Service-level unit tests: generate_report_template()
# ---------------------------------------------------------------------------

class TestGenerateReportTemplateService:
    """Unit tests for the report_template service function."""

    def test_flat_rate_contract_returns_bytes(self):
        """generate_report_template returns non-empty bytes for a flat-rate contract."""
        from app.services.report_template import generate_report_template

        contract = _make_db_contract(royalty_rate="8%")
        result = generate_report_template(contract)

        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_flat_rate_returns_valid_xlsx(self):
        """Bytes returned for a flat-rate contract can be loaded as a valid xlsx workbook."""
        import openpyxl

        from app.services.report_template import generate_report_template

        contract = _make_db_contract(royalty_rate="8%")
        result = generate_report_template(contract)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) >= 1

    def test_flat_rate_has_required_columns(self):
        """
        Flat-rate template has Period Start, Period End, Net Sales,
        Reported Royalty columns — no Category column.
        """
        import openpyxl

        from app.services.report_template import generate_report_template

        contract = _make_db_contract(royalty_rate="8%")
        result = generate_report_template(contract)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active

        # Collect all cell values from the sheet
        all_values = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    all_values.append(str(cell))

        # Required columns must appear somewhere in the sheet
        assert any("Period Start" in v for v in all_values)
        assert any("Period End" in v for v in all_values)
        assert any("Net Sales" in v for v in all_values)
        # Royalty column present
        assert any("Royalty" in v for v in all_values)

    def test_flat_rate_has_no_category_column(self):
        """Flat-rate template should NOT have a Category column."""
        import openpyxl

        from app.services.report_template import generate_report_template

        contract = _make_db_contract(royalty_rate="8%")
        result = generate_report_template(contract)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active

        all_values = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    all_values.append(str(cell).lower())

        assert not any(v == "category" for v in all_values)

    def test_category_rate_includes_category_column(self):
        """Category-rate contract template includes a Category column."""
        import openpyxl

        from app.services.report_template import generate_report_template

        contract = _make_db_contract(
            royalty_rate={"Apparel": "8%", "Accessories": "10%"}
        )
        result = generate_report_template(contract)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active

        all_values = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    all_values.append(str(cell))

        assert any("Category" in v for v in all_values)

    def test_tiered_rate_has_required_columns(self):
        """Tiered-rate contract template has same columns as flat-rate (no Category)."""
        import openpyxl

        from app.services.report_template import generate_report_template

        tiered_rate = [
            {"threshold": "$0-$2,000,000", "rate": "6%"},
            {"threshold": "$2,000,001+", "rate": "8%"},
        ]
        contract = _make_db_contract(royalty_rate=tiered_rate)
        result = generate_report_template(contract)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active

        all_values = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    all_values.append(str(cell))

        assert any("Period Start" in v for v in all_values)
        assert any("Net Sales" in v for v in all_values)
        # No Category for tiered
        assert not any(v.strip().lower() == "category" for v in all_values)

    def test_title_row_contains_licensee_name(self):
        """The generated template contains the licensee name in the title area."""
        import openpyxl

        from app.services.report_template import generate_report_template

        contract = _make_db_contract(licensee_name="Sunrise Apparel Co.")
        result = generate_report_template(contract)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active

        all_values = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    all_values.append(str(cell))

        assert any("Sunrise Apparel Co." in v for v in all_values)

    def test_header_row_is_bold(self):
        """Header row cells have bold font formatting."""
        import openpyxl

        from app.services.report_template import generate_report_template

        contract = _make_db_contract(royalty_rate="8%")
        result = generate_report_template(contract)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active

        # Find the header row: a row containing "Net Sales"
        header_row_idx = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "Net Sales" in str(cell.value):
                    header_row_idx = cell.row
                    break
            if header_row_idx is not None:
                break

        assert header_row_idx is not None, "Could not find header row"

        # Check that at least one cell in the header row is bold
        header_cells = list(ws[header_row_idx])
        bold_cells = [c for c in header_cells if c.font and c.font.bold]
        assert len(bold_cells) > 0, "No bold cells found in header row"

    def test_column_names_match_suggest_mapping_synonyms(self):
        """
        Column header names are recognized by suggest_mapping (auto-mapping
        works when the template is uploaded back).
        """
        import openpyxl

        from app.services.report_template import generate_report_template
        from app.services.spreadsheet_parser import suggest_mapping

        contract = _make_db_contract(royalty_rate="8%")
        result = generate_report_template(contract)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active

        # Find the header row (row containing "Net Sales")
        header_row = None
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and "Net Sales" in str(cell) for cell in row):
                header_row = [str(c) for c in row if c is not None]
                break

        assert header_row is not None, "Could not find header row"

        mapping = suggest_mapping(header_row, saved_mapping=None)

        # Net Sales column must be recognized
        assert any(v == "net_sales" for v in mapping.values()), (
            f"net_sales not recognized in mapping: {mapping}"
        )
        # Reported Royalty column must be recognized as licensee_reported_royalty
        assert any(v == "licensee_reported_royalty" for v in mapping.values()), (
            f"licensee_reported_royalty not recognized in mapping: {mapping}"
        )

    def test_category_column_recognized_by_suggest_mapping(self):
        """Category column in category-rate template is recognized by suggest_mapping."""
        import openpyxl

        from app.services.report_template import generate_report_template
        from app.services.spreadsheet_parser import suggest_mapping

        contract = _make_db_contract(
            royalty_rate={"Apparel": "8%", "Accessories": "10%"}
        )
        result = generate_report_template(contract)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active

        header_row = None
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and "Net Sales" in str(cell) for cell in row):
                header_row = [str(c) for c in row if c is not None]
                break

        assert header_row is not None
        mapping = suggest_mapping(header_row, saved_mapping=None)

        assert any(v == "product_category" for v in mapping.values()), (
            f"product_category not recognized in mapping: {mapping}"
        )


# ---------------------------------------------------------------------------
# Endpoint tests: GET /api/contracts/{id}/report-template
# ---------------------------------------------------------------------------

class TestReportTemplateEndpoint:
    """Integration tests for the report-template endpoint."""

    @pytest.mark.asyncio
    async def test_returns_xlsx_bytes_for_active_contract(self):
        """Endpoint returns 200 with xlsx content for an active contract."""
        import openpyxl

        from app.routers.contracts import get_report_template

        contract = _make_db_contract(royalty_rate="8%", status="active")

        with patch("app.routers.contracts.verify_contract_ownership", new_callable=AsyncMock) as mock_verify:
            # verify_contract_ownership now returns the contract row directly
            mock_verify.return_value = contract

            response = await get_report_template(
                contract_id="contract-123",
                user_id="user-123",
            )

        # StreamingResponse: read the body
        body = b"".join([chunk async for chunk in response.body_iterator])
        assert len(body) > 0

        # Must be valid xlsx
        wb = openpyxl.load_workbook(io.BytesIO(body))
        assert len(wb.sheetnames) >= 1

    @pytest.mark.asyncio
    async def test_returns_correct_content_type_header(self):
        """Endpoint sets content-type to xlsx MIME type."""
        from app.routers.contracts import get_report_template

        contract = _make_db_contract(royalty_rate="8%", status="active")

        with patch("app.routers.contracts.verify_contract_ownership", new_callable=AsyncMock) as mock_verify:
            # verify_contract_ownership now returns the contract row directly
            mock_verify.return_value = contract

            response = await get_report_template(
                contract_id="contract-123",
                user_id="user-123",
            )

        assert "spreadsheetml" in response.media_type or "openxmlformats" in response.media_type

    @pytest.mark.asyncio
    async def test_returns_content_disposition_attachment(self):
        """Endpoint sets Content-Disposition to attachment with .xlsx filename."""
        from app.routers.contracts import get_report_template

        contract = _make_db_contract(
            licensee_name="Sunrise Apparel Co.", royalty_rate="8%", status="active"
        )

        with patch("app.routers.contracts.verify_contract_ownership", new_callable=AsyncMock) as mock_verify:
            # verify_contract_ownership now returns the contract row directly
            mock_verify.return_value = contract

            response = await get_report_template(
                contract_id="contract-123",
                user_id="user-123",
            )

        disposition = response.headers.get("Content-Disposition", "")
        assert "attachment" in disposition
        assert ".xlsx" in disposition

    @pytest.mark.asyncio
    async def test_returns_404_for_missing_contract(self):
        """Endpoint returns 404 when the contract does not exist."""
        from fastapi import HTTPException

        from app.routers.contracts import get_report_template

        with patch("app.routers.contracts.verify_contract_ownership", new_callable=AsyncMock) as mock_verify:
            # verify_contract_ownership raises 404 when contract not found
            mock_verify.side_effect = HTTPException(
                status_code=404, detail="Contract not found"
            )

            with pytest.raises(HTTPException) as exc_info:
                await get_report_template(
                    contract_id="nonexistent",
                    user_id="user-123",
                )

        assert exc_info.value.status_code == 404

    @pytest.mark.asyncio
    async def test_returns_409_for_draft_contract(self):
        """Endpoint returns 409 when the contract is in draft status."""
        from fastapi import HTTPException

        from app.routers.contracts import get_report_template

        contract = _make_db_contract(status="draft")

        with patch("app.routers.contracts.verify_contract_ownership", new_callable=AsyncMock) as mock_verify:
            # verify_contract_ownership returns the draft contract row
            mock_verify.return_value = contract

            with pytest.raises(HTTPException) as exc_info:
                await get_report_template(
                    contract_id="contract-123",
                    user_id="user-123",
                )

        assert exc_info.value.status_code == 409

    @pytest.mark.asyncio
    async def test_category_contract_template_includes_category_column(self):
        """Category-rate contract produces a template with a Category column."""
        import openpyxl

        from app.routers.contracts import get_report_template

        contract = _make_db_contract(
            royalty_rate={"Apparel": "8%", "Accessories": "10%"},
            status="active",
        )

        with patch("app.routers.contracts.verify_contract_ownership", new_callable=AsyncMock) as mock_verify:
            # verify_contract_ownership returns the category contract row
            mock_verify.return_value = contract

            response = await get_report_template(
                contract_id="contract-123",
                user_id="user-123",
            )

        body = b"".join([chunk async for chunk in response.body_iterator])
        wb = openpyxl.load_workbook(io.BytesIO(body))
        ws = wb.active

        all_values = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    all_values.append(str(cell))

        assert any("Category" in v for v in all_values)
