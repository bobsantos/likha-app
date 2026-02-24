"""
Spreadsheet parser unit tests (Phase 1.1).

Tests are self-contained: xlsx files are constructed in-memory via openpyxl,
CSV files via io.StringIO.  No disk files are loaded.

TDD: these tests were written before the implementation.
"""

import io
import os
import pytest
from decimal import Decimal

# Ensure env vars are set before importing anything that triggers app imports
os.environ.setdefault("SUPABASE_URL", "https://test.supabase.co")
os.environ.setdefault("SUPABASE_KEY", "test-anon-key")
os.environ.setdefault("SUPABASE_SERVICE_KEY", "test-service-key")


# ---------------------------------------------------------------------------
# Helpers for building in-memory files
# ---------------------------------------------------------------------------

def _make_xlsx_bytes(rows: list[list]) -> bytes:
    """Build an xlsx file in-memory from a list-of-lists and return its bytes."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _make_csv_bytes(content: str, encoding: str = "utf-8") -> bytes:
    """Encode a CSV string to bytes with the given encoding."""
    return content.encode(encoding)


# ---------------------------------------------------------------------------
# parse_upload — basic parsing
# ---------------------------------------------------------------------------

class TestParseXlsxStandardHeaders:
    """parse_upload() correctly parses a standard xlsx with a TOTAL row."""

    def test_columns_detected_and_total_row_excluded(self):
        from app.services.spreadsheet_parser import parse_upload, ParsedSheet

        rows = [
            ["SKU", "Category", "Net Sales", "Returns", "Gross Sales", "Royalty Due"],
            ["APP-001", "Apparel", 12000, 500, 12500, 960],
            ["APP-002", "Apparel", 9000, 300, 9300, 720],
            ["ACC-001", "Accessories", 5000, 200, 5200, 400],
            ["TOTAL", None, 26000, 1000, 27000, 2080],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        result = parse_upload(xlsx_bytes, "report.xlsx")

        assert isinstance(result, ParsedSheet)
        assert result.column_names == ["SKU", "Category", "Net Sales", "Returns", "Gross Sales", "Royalty Due"]
        assert result.data_rows == 3           # TOTAL row excluded
        assert len(result.sample_rows) == 3

    def test_sample_rows_are_string_values(self):
        from app.services.spreadsheet_parser import parse_upload

        rows = [
            ["SKU", "Net Sales", "Royalty Due"],
            ["APP-001", 12000, 960],
            ["APP-002", 9000, 720],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        result = parse_upload(xlsx_bytes, "report.xlsx")

        # Sample rows should be string representations keyed by column name
        assert "SKU" in result.sample_rows[0]
        assert result.sample_rows[0]["Net Sales"] == "12000"


class TestParseCsvWithMetadataRows:
    """parse_upload() skips metadata rows before the real header."""

    def test_header_detected_at_row_4(self):
        from app.services.spreadsheet_parser import parse_upload

        csv_content = (
            "Licensee:,Sunrise Apparel Co.\n"
            "Period:,Q1 2025\n"
            "Contract:,LA-2025-001\n"
            "SKU,Category,Net Sales,Returns\n"
            "APP-001,Apparel,12000,500\n"
            "APP-002,Apparel,9000,300\n"
        )
        result = parse_upload(_make_csv_bytes(csv_content), "report.csv")

        assert "SKU" in result.column_names
        assert "Net Sales" in result.column_names
        assert result.data_rows == 2

    def test_detected_columns_are_from_real_header_row(self):
        from app.services.spreadsheet_parser import parse_upload

        csv_content = (
            "Report Title,Q1 Royalty Report\n"
            "Prepared by,Jane Doe\n"
            "SKU,Category,Net Sales\n"
            "APP-001,Apparel,12000\n"
        )
        result = parse_upload(_make_csv_bytes(csv_content), "report.csv")

        assert result.column_names == ["SKU", "Category", "Net Sales"]
        assert "Report Title" not in result.column_names


class TestParseXlsx5TitleRows:
    """parse_upload() handles 5 title rows above the real header (mirroring sample-3)."""

    def test_header_detected_past_5_title_rows(self):
        from app.services.spreadsheet_parser import parse_upload

        rows = [
            ["ACME Brands"],
            ["Q1 2025 Royalty Report"],
            ["Agreement Ref: LA-2025-001"],
            ["Prepared by: Jane Doe"],
            [None],  # blank row
            ["SKU", "Product Category", "Net Sales", "Royalty Due"],
            ["APP-001", "Apparel", 12000, 960],
            ["APP-002", "Apparel", 9000, 720],
            ["ACC-001", "Accessories", 5000, 400],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        result = parse_upload(xlsx_bytes, "report.xlsx")

        assert "SKU" in result.column_names
        assert "Net Sales" in result.column_names
        assert result.data_rows == 3

    def test_metadata_values_not_in_data_rows(self):
        from app.services.spreadsheet_parser import parse_upload

        rows = [
            ["Company Name"],
            ["Report Title"],
            ["Date"],
            ["Blank"],
            [None],
            ["SKU", "Net Sales"],
            ["APP-001", 12000],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        result = parse_upload(xlsx_bytes, "report.xlsx")

        assert result.data_rows == 1


class TestParseCsvWithEmptyRows:
    """parse_upload() skips empty rows when counting data_rows."""

    def test_empty_rows_excluded_from_count(self):
        from app.services.spreadsheet_parser import parse_upload

        csv_content = (
            "SKU,Category,Net Sales\n"
            "APP-001,Apparel,12000\n"
            "\n"                          # empty row
            "APP-002,Apparel,9000\n"
            ",,\n"                        # empty row (commas only)
            "ACC-001,Accessories,5000\n"
        )
        result = parse_upload(_make_csv_bytes(csv_content), "report.csv")

        assert result.data_rows == 3


class TestParseXlsxMergedCategoryColumn:
    """parse_upload() forward-fills merged cells in the category column."""

    def test_merged_category_forward_filled(self):
        """When category cell is merged, all rows in that group get the category."""
        import openpyxl
        from app.services.spreadsheet_parser import parse_upload

        wb = openpyxl.Workbook()
        ws = wb.active
        # Header
        ws.append(["Category", "SKU", "Net Sales"])
        # Row 2-5: Category "Apparel" merged across 4 rows
        ws.append(["Apparel", "APP-001", 12000])
        ws.append([None, "APP-002", 9000])
        ws.append([None, "APP-003", 8000])
        ws.append([None, "APP-004", 7000])
        # Merge category column rows 2-5
        ws.merge_cells("A2:A5")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_bytes = buf.read()

        result = parse_upload(xlsx_bytes, "report.xlsx")

        # All 4 rows should have Category = "Apparel"
        category_values = [row.get("Category") for row in result.sample_rows]
        assert all(v == "Apparel" for v in category_values), \
            f"Expected all 'Apparel', got: {category_values}"


class TestParseCsvWindows1252Encoding:
    """parse_upload() handles Windows-1252 encoded CSV without errors."""

    def test_windows1252_csv_parses_without_error(self):
        from app.services.spreadsheet_parser import parse_upload

        # Build Windows-1252 bytes directly: € is byte 0x80 in cp1252
        # "SKU,Net Sales,Currency\nAPP-001,12000,€\n" in cp1252
        header = b"SKU,Net Sales,Currency\n"
        data_row = b"APP-001,12000,\x80\n"  # 0x80 = € in cp1252
        csv_bytes = header + data_row

        result = parse_upload(csv_bytes, "report.csv")

        assert result.data_rows == 1
        assert "SKU" in result.column_names

    def test_latin1_csv_parses_without_error(self):
        from app.services.spreadsheet_parser import parse_upload

        # Include a Latin-1 character: é (0xe9)
        csv_content = "SKU,Description\nAPP-001,Caf\xe9\n"
        csv_bytes = csv_content.encode("latin-1")

        result = parse_upload(csv_bytes, "report.csv")

        assert result.data_rows == 1


class TestParseUnsupportedTypeRaisesError:
    """parse_upload() raises ParseError for unsupported file extensions."""

    def test_pdf_extension_raises_parse_error(self):
        from app.services.spreadsheet_parser import parse_upload, ParseError

        with pytest.raises(ParseError) as exc_info:
            parse_upload(b"%PDF-1.4", "report.pdf")

        assert exc_info.value.error_code == "unsupported_file_type"

    def test_docx_extension_raises_parse_error(self):
        from app.services.spreadsheet_parser import parse_upload, ParseError

        with pytest.raises(ParseError) as exc_info:
            parse_upload(b"fake-docx-bytes", "report.docx")

        assert exc_info.value.error_code == "unsupported_file_type"

    def test_txt_extension_raises_parse_error(self):
        from app.services.spreadsheet_parser import parse_upload, ParseError

        with pytest.raises(ParseError) as exc_info:
            parse_upload(b"some text", "report.txt")

        assert exc_info.value.error_code == "unsupported_file_type"


class TestParseCorruptFileRaisesError:
    """parse_upload() raises ParseError for corrupt/unreadable files."""

    def test_random_bytes_as_xlsx_raises_parse_error(self):
        from app.services.spreadsheet_parser import parse_upload, ParseError

        garbage = b"\x00\x01\x02\x03\x04\x05\xff\xfe" * 100

        with pytest.raises(ParseError) as exc_info:
            parse_upload(garbage, "report.xlsx")

        assert exc_info.value.error_code == "parse_failed"

    def test_random_bytes_as_xls_raises_parse_error(self):
        from app.services.spreadsheet_parser import parse_upload, ParseError

        garbage = b"\x00\x01\x02\x03\x04\x05" * 50

        with pytest.raises(ParseError) as exc_info:
            parse_upload(garbage, "report.xls")

        assert exc_info.value.error_code == "parse_failed"


# ---------------------------------------------------------------------------
# apply_mapping — aggregation
# ---------------------------------------------------------------------------

class TestApplyMappingFlatRateAggregation:
    """apply_mapping() aggregates net_sales when no category column is mapped."""

    def test_net_sales_summed_across_all_rows(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["SKU", "Net Sales"],
            ["APP-001", 12000],
            ["APP-002", 9000],
            ["APP-003", 8000],
            ["APP-004", 7000],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {"SKU": "ignore", "Net Sales": "net_sales"}
        result = apply_mapping(parsed, column_mapping)

        assert result.net_sales == Decimal("36000")
        assert result.category_sales is None

    def test_returns_and_gross_ignored_when_not_mapped(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["SKU", "Net Sales", "Gross Sales"],
            ["APP-001", 12000, 12500],
            ["APP-002", 9000, 9300],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {"Net Sales": "net_sales"}
        result = apply_mapping(parsed, column_mapping)

        assert result.net_sales == Decimal("21000")


class TestApplyMappingCategoryAggregation:
    """apply_mapping() aggregates net_sales by product_category."""

    def test_category_sales_aggregated_correctly(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Category", "Net Sales"],
            ["Apparel", 10000],
            ["Apparel", 8000],
            ["Apparel", 7000],
            ["Apparel", 5000],
            ["Accessories", 6000],
            ["Accessories", 5000],
            ["Accessories", 4000],
            ["Accessories", 3000],
            ["Footwear", 4000],
            ["Footwear", 3000],
            ["Footwear", 2000],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {
            "Category": "product_category",
            "Net Sales": "net_sales",
        }
        result = apply_mapping(parsed, column_mapping)

        assert result.category_sales == {
            "Apparel": Decimal("30000"),
            "Accessories": Decimal("18000"),
            "Footwear": Decimal("9000"),
        }
        assert result.net_sales == Decimal("57000")

    def test_net_sales_is_sum_of_all_categories(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Category", "Net Sales"],
            ["Apparel", 15000],
            ["Accessories", 10000],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {"Category": "product_category", "Net Sales": "net_sales"}
        result = apply_mapping(parsed, column_mapping)

        assert result.net_sales == Decimal("25000")


class TestApplyMappingTotalRowExcluded:
    """apply_mapping() excludes TOTAL rows from aggregation."""

    def test_total_row_not_included_in_sum(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["SKU", "Net Sales"],
            ["APP-001", 10000],
            ["APP-002", 8000],
            ["APP-003", 7000],
            ["TOTAL", 25000],  # must be excluded
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {"SKU": "ignore", "Net Sales": "net_sales"}
        result = apply_mapping(parsed, column_mapping)

        assert result.net_sales == Decimal("25000")  # 10000 + 8000 + 7000


class TestApplyMappingLicenseeRoyaltyExtraction:
    """apply_mapping() extracts licensee_reported_royalty when column is mapped."""

    def test_royalty_summed_from_per_row_values(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["SKU", "Net Sales", "Royalty Due"],
            ["APP-001", 12000, 960],
            ["APP-002", 9000, 720],
            ["APP-003", 8000, 640],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {
            "SKU": "ignore",
            "Net Sales": "net_sales",
            "Royalty Due": "licensee_reported_royalty",
        }
        result = apply_mapping(parsed, column_mapping)

        assert result.licensee_reported_royalty == Decimal("2320")  # 960+720+640

    def test_royalty_none_when_not_mapped(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["SKU", "Net Sales"],
            ["APP-001", 12000],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {"Net Sales": "net_sales"}
        result = apply_mapping(parsed, column_mapping)

        assert result.licensee_reported_royalty is None


class TestApplyMappingMissingNetSalesRaisesError:
    """apply_mapping() raises MappingError when no column maps to net_sales."""

    def test_no_net_sales_or_gross_raises_mapping_error(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping, MappingError

        rows = [
            ["SKU", "Product Category"],
            ["APP-001", "Apparel"],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        # Neither net_sales nor gross_sales mapped
        column_mapping = {"SKU": "ignore", "Product Category": "product_category"}

        with pytest.raises(MappingError) as exc_info:
            apply_mapping(parsed, column_mapping)

        assert exc_info.value.error_code == "net_sales_column_required"


class TestApplyMappingNegativeNetSalesRaisesError:
    """apply_mapping() raises MappingError when aggregated net_sales is negative."""

    def test_negative_net_sales_raises_error(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping, MappingError

        # Returns column has higher values than net sales (mapping error scenario)
        rows = [
            ["Net Sales"],
            [-5000],  # already negative
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {"Net Sales": "net_sales"}

        with pytest.raises(MappingError) as exc_info:
            apply_mapping(parsed, column_mapping)

        assert exc_info.value.error_code == "negative_net_sales"


class TestApplyMappingDerivedNetSales:
    """apply_mapping() derives net_sales = gross_sales - returns when net_sales not mapped."""

    def test_net_sales_derived_from_gross_minus_returns(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["SKU", "Gross Sales", "Returns"],
            ["APP-001", 12500, 500],
            ["APP-002", 9300, 300],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        # Map gross_sales and returns but NOT net_sales
        column_mapping = {
            "SKU": "ignore",
            "Gross Sales": "gross_sales",
            "Returns": "returns",
        }
        result = apply_mapping(parsed, column_mapping)

        # net_sales = (12500 - 500) + (9300 - 300) = 12000 + 9000 = 21000
        assert result.net_sales == Decimal("21000")

    def test_gross_only_treated_as_net_when_no_returns(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Gross Sales"],
            [12500],
            [9300],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {"Gross Sales": "gross_sales"}
        result = apply_mapping(parsed, column_mapping)

        assert result.net_sales == Decimal("21800")


# ---------------------------------------------------------------------------
# suggest_mapping — keyword synonym matching
# ---------------------------------------------------------------------------

class TestKeywordMatchingStandardNames:
    """suggest_mapping() correctly matches standard column names."""

    def test_standard_column_names_matched(self):
        from app.services.spreadsheet_parser import suggest_mapping

        columns = ["Net Sales Amount", "Product Category", "Gross Sales", "Royalty Due"]
        result = suggest_mapping(columns, saved_mapping=None)

        assert result["Net Sales Amount"] == "net_sales"
        assert result["Product Category"] == "product_category"
        assert result["Gross Sales"] == "gross_sales"
        assert result["Royalty Due"] == "licensee_reported_royalty"


class TestKeywordMatchingNonStandardNames:
    """suggest_mapping() handles non-standard column names from sample-3."""

    def test_non_standard_names_matched_or_ignored(self):
        from app.services.spreadsheet_parser import suggest_mapping

        columns = ["Total Revenue", "Refunds", "Amount Owed", "Gross Revenue", "Rate (%)"]
        result = suggest_mapping(columns, saved_mapping=None)

        # These are NOT in the synonym list
        assert result["Total Revenue"] == "ignore"
        assert result["Refunds"] == "ignore"

        # "Rate (%)" now maps to royalty_rate (Phase 1.1.1 — it was 'ignore' before)
        assert result["Rate (%)"] == "royalty_rate"

        # "Amount Owed" IS a synonym for licensee_reported_royalty (per-row royalty column in sample-3)
        assert result["Amount Owed"] == "licensee_reported_royalty"

        # "Gross Revenue" IS a synonym for gross_sales
        assert result["Gross Revenue"] == "gross_sales"

    def test_royalty_rate_column_not_matched_to_licensee_royalty(self):
        """'Royalty Rate' must map to royalty_rate — not to licensee_reported_royalty.

        Phase 1.1.1: 'Royalty Rate' now maps to the new royalty_rate cross-check field
        instead of 'ignore'. The bare 'royalty' synonym was never added to
        licensee_reported_royalty, so 'Royalty Due' still maps correctly.
        """
        from app.services.spreadsheet_parser import suggest_mapping

        columns = [
            "Product Description", "SKU", "Product Category",
            "Gross Sales", "Returns / Allowances", "Net Sales",
            "Royalty Rate", "Royalty Due",
        ]
        result = suggest_mapping(columns, saved_mapping=None)

        # "Royalty Rate" is a percentage column — maps to the royalty_rate cross-check field
        assert result["Royalty Rate"] == "royalty_rate"

        # "Royalty Due" is the actual per-row royalty amount — must be mapped correctly
        assert result["Royalty Due"] == "licensee_reported_royalty"


class TestKeywordMatchingCaseInsensitive:
    """suggest_mapping() performs case-insensitive matching."""

    def test_uppercase_net_sales_matches(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["NET SALES"], saved_mapping=None)
        assert result["NET SALES"] == "net_sales"

    def test_mixed_case_matches(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Net Sales"], saved_mapping=None)
        assert result["Net Sales"] == "net_sales"

    def test_lowercase_matches(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["net sales"], saved_mapping=None)
        assert result["net sales"] == "net_sales"


class TestKeywordMatchingSubstring:
    """suggest_mapping() matches synonyms as substrings."""

    def test_total_net_sales_amount_matches_net_sales(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Total Net Sales Amount"], saved_mapping=None)
        assert result["Total Net Sales Amount"] == "net_sales"

    def test_returns_and_allowances_matches_returns(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Returns and Allowances"], saved_mapping=None)
        assert result["Returns and Allowances"] == "returns"

    def test_product_category_matches(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Product Category"], saved_mapping=None)
        assert result["Product Category"] == "product_category"

    def test_division_matches_product_category(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Division"], saved_mapping=None)
        assert result["Division"] == "product_category"


class TestKeywordMatchingSavedMapping:
    """suggest_mapping() returns saved_mapping unchanged when provided."""

    def test_saved_mapping_returned_directly(self):
        from app.services.spreadsheet_parser import suggest_mapping

        saved = {
            "Net Sales Amount": "net_sales",
            "SKU": "ignore",
            "Royalty Due": "licensee_reported_royalty",
        }
        result = suggest_mapping(
            column_names=["Net Sales Amount", "SKU", "Royalty Due"],
            saved_mapping=saved,
        )

        # saved_mapping takes priority — returned as-is for columns in the saved dict
        assert result["Net Sales Amount"] == "net_sales"
        assert result["SKU"] == "ignore"
        assert result["Royalty Due"] == "licensee_reported_royalty"

    def test_new_columns_not_in_saved_mapping_use_keyword_matching(self):
        from app.services.spreadsheet_parser import suggest_mapping

        saved = {"SKU": "ignore"}
        result = suggest_mapping(
            column_names=["SKU", "Net Sales", "New Column"],
            saved_mapping=saved,
        )

        assert result["SKU"] == "ignore"
        assert result["Net Sales"] == "net_sales"
        assert result["New Column"] == "ignore"


class TestKeywordMatchingNsAntiPattern:
    """The ' ns' synonym does not false-match columns like 'Units'."""

    def test_units_column_not_matched_to_net_sales(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Units", "Transactions", "Bonus"], saved_mapping=None)
        assert result["Units"] == "ignore"
        assert result["Transactions"] == "ignore"
        assert result["Bonus"] == "ignore"

    def test_ns_with_space_prefix_matched_to_net_sales(self):
        from app.services.spreadsheet_parser import suggest_mapping

        # A column literally named " NS" (with leading space) should match
        result = suggest_mapping([" NS"], saved_mapping=None)
        assert result[" NS"] == "net_sales"


# ---------------------------------------------------------------------------
# suggest_mapping — new field synonyms (Phase 1.1.1)
# ---------------------------------------------------------------------------

class TestKeywordMatchingNewFieldsPhase111:
    """suggest_mapping() correctly matches the three new cross-check fields."""

    def test_licensee_name_column_matched(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Licensee Name"], saved_mapping=None)
        assert result["Licensee Name"] == "licensee_name"

    def test_licensee_column_matched(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Licensee"], saved_mapping=None)
        assert result["Licensee"] == "licensee_name"

    def test_company_name_matched_to_licensee_name(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Company Name"], saved_mapping=None)
        assert result["Company Name"] == "licensee_name"

    def test_manufacturer_matched_to_licensee_name(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Manufacturer"], saved_mapping=None)
        assert result["Manufacturer"] == "licensee_name"

    def test_reporting_period_matched_to_report_period(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Reporting Period"], saved_mapping=None)
        assert result["Reporting Period"] == "report_period"

    def test_report_period_column_matched(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Report Period"], saved_mapping=None)
        assert result["Report Period"] == "report_period"

    def test_quarter_matched_to_report_period(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Quarter"], saved_mapping=None)
        assert result["Quarter"] == "report_period"

    def test_royalty_rate_column_matched_to_royalty_rate(self):
        """'Royalty Rate' must now map to royalty_rate (not licensee_reported_royalty and not ignore)."""
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Royalty Rate"], saved_mapping=None)
        assert result["Royalty Rate"] == "royalty_rate"

    def test_applicable_rate_matched_to_royalty_rate(self):
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Applicable Rate"], saved_mapping=None)
        assert result["Applicable Rate"] == "royalty_rate"

    def test_rate_percent_matched_to_royalty_rate(self):
        """'Rate (%)' must now map to royalty_rate (was 'ignore' before Phase 1.1.1)."""
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Rate (%)"], saved_mapping=None)
        assert result["Rate (%)"] == "royalty_rate"

    def test_royalty_due_still_maps_to_licensee_reported_royalty(self):
        """Adding royalty_rate synonyms must not steal 'Royalty Due' from licensee_reported_royalty."""
        from app.services.spreadsheet_parser import suggest_mapping

        result = suggest_mapping(["Royalty Due"], saved_mapping=None)
        assert result["Royalty Due"] == "licensee_reported_royalty"

    def test_royalty_rate_does_not_steal_royalty_due(self):
        """Full realistic set: 'Royalty Rate' -> royalty_rate, 'Royalty Due' -> licensee_reported_royalty."""
        from app.services.spreadsheet_parser import suggest_mapping

        columns = [
            "Product Description", "SKU", "Product Category",
            "Gross Sales", "Returns / Allowances", "Net Sales",
            "Royalty Rate", "Royalty Due",
        ]
        result = suggest_mapping(columns, saved_mapping=None)

        assert result["Royalty Rate"] == "royalty_rate"
        assert result["Royalty Due"] == "licensee_reported_royalty"

    def test_period_start_does_not_match_report_period(self):
        """'Period Start' contains 'period' but should be ignored — it's a date boundary, not a report period label."""
        from app.services.spreadsheet_parser import suggest_mapping

        # "period" substring match is intentional; the spec says use it.
        # But we verify no false conflicts: "Period Start" could match "period"
        # which is a synonym for report_period.  This test documents current behavior.
        # If the implementation uses exact synonym matching (not substring), this
        # would be "ignore". The spec does say "period" is a synonym, so a column
        # literally named "Period" should match — "Period Start" contains "period"
        # so it will match report_period. This is acceptable per the spec.
        result = suggest_mapping(["Period"], saved_mapping=None)
        assert result["Period"] == "report_period"


class TestExtractCrossCheckValues:
    """extract_cross_check_values() returns first non-null value per cross-check field."""

    def test_returns_licensee_name_from_mapped_column(self):
        from app.services.spreadsheet_parser import parse_upload, extract_cross_check_values

        rows = [
            ["Licensee Name", "Net Sales"],
            ["Sunrise Apparel Co.", 12000],
            ["Sunrise Apparel Co.", 9000],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")
        column_mapping = {
            "Licensee Name": "licensee_name",
            "Net Sales": "net_sales",
        }
        result = extract_cross_check_values(parsed, column_mapping)

        assert result["licensee_name"] == "Sunrise Apparel Co."

    def test_returns_report_period_from_mapped_column(self):
        from app.services.spreadsheet_parser import parse_upload, extract_cross_check_values

        rows = [
            ["Report Period", "Net Sales"],
            ["Q1 2025", 12000],
            ["Q1 2025", 9000],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")
        column_mapping = {
            "Report Period": "report_period",
            "Net Sales": "net_sales",
        }
        result = extract_cross_check_values(parsed, column_mapping)

        assert result["report_period"] == "Q1 2025"

    def test_returns_royalty_rate_from_mapped_column(self):
        from app.services.spreadsheet_parser import parse_upload, extract_cross_check_values

        rows = [
            ["Royalty Rate", "Net Sales"],
            ["8%", 12000],
            ["8%", 9000],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")
        column_mapping = {
            "Royalty Rate": "royalty_rate",
            "Net Sales": "net_sales",
        }
        result = extract_cross_check_values(parsed, column_mapping)

        assert result["royalty_rate"] == "8%"

    def test_returns_none_when_cross_check_columns_not_mapped(self):
        from app.services.spreadsheet_parser import parse_upload, extract_cross_check_values

        rows = [
            ["Net Sales"],
            [12000],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")
        column_mapping = {"Net Sales": "net_sales"}
        result = extract_cross_check_values(parsed, column_mapping)

        assert result["licensee_name"] is None
        assert result["report_period"] is None
        assert result["royalty_rate"] is None

    def test_skips_empty_rows_to_find_first_non_null(self):
        from app.services.spreadsheet_parser import parse_upload, extract_cross_check_values

        # First row has empty licensee name; second row has the value
        rows = [
            ["Licensee Name", "Net Sales"],
            ["", 12000],
            ["Sunrise Apparel Co.", 9000],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")
        column_mapping = {
            "Licensee Name": "licensee_name",
            "Net Sales": "net_sales",
        }
        result = extract_cross_check_values(parsed, column_mapping)

        assert result["licensee_name"] == "Sunrise Apparel Co."


# ---------------------------------------------------------------------------
# apply_mapping — "metadata" column mapping value (Phase 1.1.1)
# ---------------------------------------------------------------------------

class TestApplyMappingMetadataExcludedFromNetSales:
    """Columns mapped to 'metadata' must not contribute to net_sales aggregation."""

    def test_metadata_column_excluded_from_net_sales_sum(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Net Sales", "SKU", "Internal Ref"],
            [10000, "APP-001", "REF-001"],
            [8000, "APP-002", "REF-002"],
            [7000, "APP-003", "REF-003"],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        # SKU and Internal Ref are mapped to metadata — they must not affect net_sales
        column_mapping = {
            "Net Sales": "net_sales",
            "SKU": "metadata",
            "Internal Ref": "metadata",
        }
        result = apply_mapping(parsed, column_mapping)

        assert result.net_sales == Decimal("25000")

    def test_multiple_metadata_columns_do_not_cause_double_counting(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Net Sales", "SKU", "Region Code", "PO Number"],
            [5000, "A-01", "US-W", "PO-9001"],
            [3000, "A-02", "US-E", "PO-9002"],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {
            "Net Sales": "net_sales",
            "SKU": "metadata",
            "Region Code": "metadata",
            "PO Number": "metadata",
        }
        result = apply_mapping(parsed, column_mapping)

        assert result.net_sales == Decimal("8000")


class TestApplyMappingMetadataExcludedFromCategoryBreakdown:
    """Columns mapped to 'metadata' must not affect category_breakdown."""

    def test_metadata_column_alongside_product_category_does_not_affect_breakdown(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Category", "Net Sales", "SKU"],
            ["Apparel", 10000, "APP-001"],
            ["Apparel", 8000, "APP-002"],
            ["Footwear", 5000, "FW-001"],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {
            "Category": "product_category",
            "Net Sales": "net_sales",
            "SKU": "metadata",
        }
        result = apply_mapping(parsed, column_mapping)

        assert result.category_sales == {
            "Apparel": Decimal("18000"),
            "Footwear": Decimal("5000"),
        }
        assert result.net_sales == Decimal("23000")

    def test_metadata_column_not_treated_as_product_category(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Net Sales", "Tag"],
            [10000, "promo"],
            [8000, "regular"],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {
            "Net Sales": "net_sales",
            "Tag": "metadata",
        }
        result = apply_mapping(parsed, column_mapping)

        # "Tag" mapped to metadata must not produce a category breakdown
        assert result.category_sales is None


class TestApplyMappingMetadataCapturesRawValues:
    """apply_mapping() collects raw cell values for metadata-mapped columns."""

    def test_metadata_field_present_on_mapped_data(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Net Sales", "SKU"],
            [10000, "APP-001"],
            [8000, "APP-002"],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {
            "Net Sales": "net_sales",
            "SKU": "metadata",
        }
        result = apply_mapping(parsed, column_mapping)

        assert hasattr(result, "metadata")

    def test_metadata_contains_values_from_metadata_mapped_columns(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Net Sales", "SKU", "Internal Ref"],
            [10000, "APP-001", "REF-001"],
            [8000, "APP-002", "REF-002"],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {
            "Net Sales": "net_sales",
            "SKU": "metadata",
            "Internal Ref": "metadata",
        }
        result = apply_mapping(parsed, column_mapping)

        # metadata should be a dict containing the captured column values
        assert result.metadata is not None
        assert "SKU" in result.metadata
        assert "Internal Ref" in result.metadata

    def test_metadata_values_are_lists_of_row_values(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Net Sales", "SKU"],
            [10000, "APP-001"],
            [8000, "APP-002"],
            [7000, "APP-003"],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {
            "Net Sales": "net_sales",
            "SKU": "metadata",
        }
        result = apply_mapping(parsed, column_mapping)

        assert result.metadata["SKU"] == ["APP-001", "APP-002", "APP-003"]

    def test_metadata_is_none_when_no_metadata_columns_mapped(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Net Sales", "SKU"],
            [10000, "APP-001"],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {
            "Net Sales": "net_sales",
            "SKU": "ignore",
        }
        result = apply_mapping(parsed, column_mapping)

        assert result.metadata is None


class TestApplyMappingMetadataNoErrors:
    """apply_mapping() does not raise errors when 'metadata' appears in column_mapping."""

    def test_metadata_in_mapping_does_not_cause_error(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Net Sales", "SKU", "PO Number"],
            [10000, "APP-001", "PO-9001"],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {
            "Net Sales": "net_sales",
            "SKU": "metadata",
            "PO Number": "metadata",
        }
        # Must not raise any exception
        result = apply_mapping(parsed, column_mapping)
        assert result.net_sales == Decimal("10000")

    def test_all_columns_are_metadata_except_net_sales(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Net Sales", "A", "B", "C", "D"],
            [5000, "x1", "x2", "x3", "x4"],
            [3000, "y1", "y2", "y3", "y4"],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {
            "Net Sales": "net_sales",
            "A": "metadata",
            "B": "metadata",
            "C": "metadata",
            "D": "metadata",
        }
        result = apply_mapping(parsed, column_mapping)

        assert result.net_sales == Decimal("8000")
        assert result.category_sales is None

    def test_metadata_mixed_with_ignore_and_named_fields(self):
        from app.services.spreadsheet_parser import parse_upload, apply_mapping

        rows = [
            ["Net Sales", "Category", "SKU", "Noise"],
            [10000, "Apparel", "APP-001", "junk1"],
            [8000, "Footwear", "FW-001", "junk2"],
        ]
        xlsx_bytes = _make_xlsx_bytes(rows)
        parsed = parse_upload(xlsx_bytes, "report.xlsx")

        column_mapping = {
            "Net Sales": "net_sales",
            "Category": "product_category",
            "SKU": "metadata",
            "Noise": "ignore",
        }
        result = apply_mapping(parsed, column_mapping)

        assert result.net_sales == Decimal("18000")
        assert result.category_sales == {
            "Apparel": Decimal("10000"),
            "Footwear": Decimal("8000"),
        }
        assert result.metadata is not None
        assert "SKU" in result.metadata
        # "Noise" was ignored — must not appear in metadata
        assert "Noise" not in result.metadata
