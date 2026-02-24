"""
Report template generation service (Phase 2, Task 1).

Generates a pre-formatted Excel (.xlsx) file that a licensor can email to
their licensee. The licensee fills in the spreadsheet and returns it. When
the licensor uploads the returned file, column headers match exactly what
spreadsheet_parser.suggest_mapping recognizes, enabling zero-effort mapping.

Public API:
  generate_report_template(contract: dict) -> bytes
"""

import io
import logging
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# These names are chosen to match synonyms in spreadsheet_parser.FIELD_SYNONYMS
# so that suggest_mapping auto-recognizes them when the file is uploaded back.
#
# Mapping targets:
#   "Period Start"       → report_period   (via "period" synonym)
#   "Period End"         → report_period   (via "period" synonym)
#   "Category"           → product_category (via "category" synonym)
#   "Net Sales"          → net_sales       (via "net sales" synonym)
#   "Reported Royalty"   → licensee_reported_royalty (via "royalty due"? No—use "total royalty")
#
# Note: "Reported Royalty" doesn't match any existing synonym directly.
# We use "Royalty Due" which IS a synonym for licensee_reported_royalty.
FLAT_COLUMNS = [
    "Period Start",
    "Period End",
    "Net Sales",
    "Royalty Due",
]

CATEGORY_COLUMNS = [
    "Period Start",
    "Period End",
    "Category",
    "Net Sales",
    "Royalty Due",
]

# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=13)
_SUBTITLE_FONT = Font(bold=False, size=10, italic=True)

_HEADER_FILL = PatternFill(
    start_color="D9E1F2",
    end_color="D9E1F2",
    fill_type="solid",
)

# Column widths (characters) keyed by column header name
_COLUMN_WIDTHS: dict[str, int] = {
    "Period Start": 14,
    "Period End": 14,
    "Category": 20,
    "Net Sales": 16,
    "Royalty Due": 16,
}
_DEFAULT_COLUMN_WIDTH = 16


def _is_category_rate(royalty_rate: Any) -> bool:
    """Return True if the contract uses category-specific rates (a dict)."""
    return isinstance(royalty_rate, dict)


def _is_tiered_rate(royalty_rate: Any) -> bool:
    """Return True if the contract uses tiered rates (a list)."""
    return isinstance(royalty_rate, list)


def _rate_description(royalty_rate: Any) -> str:
    """Return a human-readable description of the royalty rate for the template."""
    if royalty_rate is None:
        return "See contract"
    if isinstance(royalty_rate, str):
        return royalty_rate
    if isinstance(royalty_rate, list):
        # Tiered: e.g. "Tiered: 6% / 8%"
        rates = [t.get("rate", "?") if isinstance(t, dict) else str(t) for t in royalty_rate]
        return "Tiered: " + " / ".join(rates)
    if isinstance(royalty_rate, dict):
        # Category: e.g. "Apparel: 8%, Accessories: 10%"
        parts = [f"{cat}: {rate}" for cat, rate in royalty_rate.items()]
        return ", ".join(parts)
    return str(royalty_rate)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_report_template(contract: dict) -> bytes:
    """
    Generate a pre-formatted Excel report template for the given contract.

    The template is designed to be emailed to the licensee who fills it in
    and returns it. Column headers use names that spreadsheet_parser's
    suggest_mapping function will auto-recognize, enabling zero-effort
    column mapping on upload.

    Args:
        contract: Contract dict from the database (must be active).

    Returns:
        Bytes of the generated .xlsx file.
    """
    royalty_rate = contract.get("royalty_rate")
    licensee_name = contract.get("licensee_name") or "Licensee"
    contract_id = contract.get("id", "")
    start_date = contract.get("contract_start_date", "")
    end_date = contract.get("contract_end_date", "")
    reporting_frequency = contract.get("reporting_frequency", "quarterly")

    use_category = _is_category_rate(royalty_rate)
    columns = CATEGORY_COLUMNS if use_category else FLAT_COLUMNS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Royalty Report"

    # ------------------------------------------------------------------
    # Row 1: Title
    # ------------------------------------------------------------------
    ws.append([f"Royalty Report — {licensee_name}"])
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = _TITLE_FONT

    # ------------------------------------------------------------------
    # Row 2: Contract period info
    # ------------------------------------------------------------------
    period_info = f"Contract period: {start_date} to {end_date}"
    if reporting_frequency:
        period_info += f"  |  Reporting: {reporting_frequency}"
    ws.append([period_info])
    ws.cell(row=2, column=1).font = _SUBTITLE_FONT

    # ------------------------------------------------------------------
    # Row 3: Rate info
    # ------------------------------------------------------------------
    rate_desc = _rate_description(royalty_rate)
    ws.append([f"Royalty rate: {rate_desc}"])
    ws.cell(row=3, column=1).font = _SUBTITLE_FONT

    # ------------------------------------------------------------------
    # Row 4: Blank separator
    # ------------------------------------------------------------------
    ws.append([])

    # ------------------------------------------------------------------
    # Row 5: Column headers
    # ------------------------------------------------------------------
    ws.append(columns)
    header_row_idx = 5

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # ------------------------------------------------------------------
    # Rows 6–15: Empty data rows (10 blank rows for the licensee to fill)
    # ------------------------------------------------------------------
    for _ in range(10):
        ws.append([""] * len(columns))

    # ------------------------------------------------------------------
    # Set column widths
    # ------------------------------------------------------------------
    for col_idx, col_name in enumerate(columns, start=1):
        width = _COLUMN_WIDTHS.get(col_name, _DEFAULT_COLUMN_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Widen column A for the title / subtitle rows
    if ws.column_dimensions["A"].width < 40:
        ws.column_dimensions["A"].width = max(
            ws.column_dimensions["A"].width, 40
        )

    # ------------------------------------------------------------------
    # Freeze header row
    # ------------------------------------------------------------------
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    # ------------------------------------------------------------------
    # Serialize to bytes
    # ------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
