"""
Spreadsheet parser service (Phase 1.1).

Parses .xlsx, .xls, and .csv files uploaded by licensors and maps detected
columns to canonical Likha field names.

Public API:
  parse_upload(file_content, filename)  -> ParsedSheet
  apply_mapping(parsed, column_mapping) -> MappedData
  suggest_mapping(column_names, saved_mapping) -> dict[str, str]
"""

import io
import json
import logging
import os
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class ParseError(Exception):
    """Raised when a file cannot be parsed."""
    def __init__(self, message: str, error_code: str):
        super().__init__(message)
        self.error_code = error_code
        self.message = message


class MappingError(Exception):
    """Raised when a column mapping is invalid or produces bad data."""
    def __init__(self, message: str, error_code: str):
        super().__init__(message)
        self.error_code = error_code
        self.message = message


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class ParsedSheet:
    """Result of parse_upload()."""
    column_names: list[str]           # detected column headers
    all_rows: list[dict]              # all data rows as dicts {col_name: str_value}
    sample_rows: list[dict]           # first 5 data rows (same format)
    data_rows: int                    # total data rows (excl. header + summary rows)
    sheet_name: str = "Sheet1"
    total_rows: int = 0               # total rows in file including header
    metadata_period_start: Optional[str] = None  # period start extracted from file metadata rows
    metadata_period_end: Optional[str] = None    # period end extracted from file metadata rows


@dataclass
class MappedData:
    """Result of apply_mapping()."""
    net_sales: Decimal
    category_sales: Optional[dict[str, Decimal]]
    licensee_reported_royalty: Optional[Decimal]
    gross_sales: Optional[Decimal] = None
    returns: Optional[Decimal] = None
    metadata: Optional[dict[str, list[str]]] = None


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

# Summary row detection keywords (first non-empty cell, case-insensitive)
SUMMARY_KEYWORDS = {"total", "subtotal", "sum", "grand total", "totals"}

# Valid canonical Likha field names
VALID_FIELDS = {
    "net_sales",
    "gross_sales",
    "returns",
    "product_category",
    "licensee_reported_royalty",
    "territory",
    "licensee_name",
    "report_period",
    "royalty_rate",
    "metadata",
    "ignore",
}

# Cross-check field names (informational — not used in calculation)
CROSS_CHECK_FIELDS = {"licensee_name", "report_period", "royalty_rate"}

# Keyword synonyms for column matching — order determines priority.
# More specific entries must appear before broader ones to prevent false matches.
# royalty_rate must come before licensee_reported_royalty so "royalty rate"
# (which contains "rate") maps to royalty_rate rather than firing on a future
# broad "royalty" synonym.
FIELD_SYNONYMS: dict[str, list[str]] = {
    "net_sales": [
        "net sales", "net revenue", "net proceeds", "royalty base",
        "net sales amount", "total net sales", " ns"
    ],
    "gross_sales": [
        "gross sales", "gross revenue", "gross proceeds", "gross amount",
        "total sales"
    ],
    "returns": [
        "returns", "allowances", "deductions", "credits",
        "returns and allowances", "r&a"
    ],
    "product_category": [
        "category", "product line", "product type", "line",
        "division", "collection", "segment"
    ],
    # royalty_rate must be checked before licensee_reported_royalty so that
    # "Royalty Rate" matches royalty_rate (via "royalty rate" or "rate")
    # instead of accidentally matching licensee_reported_royalty.
    "royalty_rate": [
        "royalty rate", "applicable rate", "rate (%)", "rate applied", "rate",
    ],
    "licensee_reported_royalty": [
        "royalty due", "amount due", "calculated royalty",
        "total royalty", "amount owed",
    ],
    "territory": [
        "territory", "region", "market", "country", "geography"
    ],
    "licensee_name": [
        "licensee name", "licensee", "company name", "company",
        "manufacturer", "partner",
    ],
    "report_period": [
        "reporting period", "report period", "fiscal period",
        "report date", "period covered", "quarter", "period",
    ],
}


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _get_extension(filename: str) -> str:
    """Return lower-case file extension including the dot."""
    dot = filename.rfind(".")
    if dot == -1:
        return ""
    return filename[dot:].lower()


def _is_numeric_value(value) -> bool:
    """Return True if value can be interpreted as a number (int, float, Decimal, or numeric string)."""
    if isinstance(value, (int, float)):
        return True
    if value is None:
        return False
    try:
        Decimal(str(value).replace(",", ""))
        return True
    except InvalidOperation:
        return False


def _is_date_like(value) -> bool:
    """Return True if value looks like a date."""
    import re
    if value is None:
        return False
    s = str(value).strip()
    # Match common date patterns: YYYY-MM-DD, MM/DD/YYYY, etc.
    date_patterns = [
        r"^\d{4}-\d{2}-\d{2}$",
        r"^\d{2}/\d{2}/\d{4}$",
        r"^\d{1,2}/\d{1,2}/\d{2,4}$",
    ]
    return any(re.match(p, s) for p in date_patterns)


def _cell_is_string_like(value) -> bool:
    """Return True if value is a non-empty string that is not numeric or date-like."""
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    if _is_numeric_value(s):
        return False
    if _is_date_like(s):
        return False
    return True


def _is_summary_row(row_cells: list) -> bool:
    """Return True if this row looks like a TOTAL/SUBTOTAL summary row."""
    for cell in row_cells:
        if cell is None:
            continue
        s = str(cell).strip().lower()
        if s in SUMMARY_KEYWORDS:
            return True
        # Also check if cell starts with a summary keyword
        for kw in SUMMARY_KEYWORDS:
            if s.startswith(kw):
                return True
        # Stop at first non-empty cell
        break
    return False


def _row_is_all_empty(row_cells: list) -> bool:
    """Return True if all cells in the row are None or empty string."""
    return all(
        cell is None or str(cell).strip() == ""
        for cell in row_cells
    )


def _looks_like_metadata_row(row: list) -> bool:
    """
    Return True if this row looks like a label:value metadata row.

    Metadata rows typically have only 1-2 populated cells where the first cell
    ends with ':' or contains a colon (a label) and the second is a value.
    """
    non_empty = [cell for cell in row if cell is not None and str(cell).strip() != ""]
    if len(non_empty) > 2:
        return False
    if not non_empty:
        return False
    first = str(non_empty[0]).strip()
    # If first cell ends with ':' it's clearly a label:value row
    if first.endswith(":"):
        return True
    # If first cell contains ':' (e.g., "Prepared by: Jane Doe" in same cell)
    if ":" in first and len(non_empty) == 1:
        return True
    return False


def _detect_header_row(all_rows: list[list]) -> int:
    """
    Find the index (0-based) of the header row.

    Algorithm:
    1. Check up to the first 20 rows.
    2. Score each row by (string_count, subsequent_numeric_count).
    3. Header row = first row with max string_count AND at least 1 subsequent
       row with numeric values.  Among ties, pick the earliest row.
    4. Skip metadata rows (label:value pairs).
    5. Fallback: first row with >= 2 string cells followed by a numeric row,
       or row 0.
    """
    max_scan = min(20, len(all_rows))

    # Calculate the max number of non-empty columns in any row
    max_cols = max(
        (sum(1 for c in row if c is not None and str(c).strip() != "")
         for row in all_rows[:max_scan] if row),
        default=1,
    )

    # Score each candidate: (string_count, index)
    candidates: list[tuple[int, int]] = []  # (string_count, row_index)

    for i in range(max_scan):
        row = all_rows[i]
        if _row_is_all_empty(row):
            continue

        # Skip metadata rows (label: value pairs with <= 2 cells)
        if _looks_like_metadata_row(row):
            continue

        string_count = sum(1 for cell in row if _cell_is_string_like(cell))
        if string_count < 2:
            continue

        # Check if there is at least one subsequent data row with numeric values
        has_numeric_following = False
        for j in range(i + 1, min(i + 6, len(all_rows))):
            next_row = all_rows[j]
            if _row_is_all_empty(next_row):
                continue
            numeric_count = sum(1 for cell in next_row if _is_numeric_value(cell))
            if numeric_count >= 1:
                has_numeric_following = True
                break

        if has_numeric_following:
            candidates.append((string_count, i))

    if not candidates:
        return 0

    # Pick the candidate with the highest string_count; on tie, pick earliest
    best_string_count = max(c[0] for c in candidates)
    for string_count, row_idx in candidates:
        if string_count == best_string_count:
            return row_idx

    return 0


def _forward_fill_column(rows: list[list], col_idx: int) -> list[list]:
    """
    Forward-fill None values in a specific column index.
    Used for merged category cells.
    """
    last_value = None
    result = []
    for row in rows:
        row_copy = list(row)
        if col_idx < len(row_copy):
            if row_copy[col_idx] is None and last_value is not None:
                row_copy[col_idx] = last_value
            elif row_copy[col_idx] is not None:
                last_value = row_copy[col_idx]
        result.append(row_copy)
    return result


def _to_decimal_safe(value) -> Optional[Decimal]:
    """Convert a value to Decimal, returning None if not possible."""
    if value is None:
        return None
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _cell_to_str(value) -> str:
    """Convert a cell value to string representation."""
    if value is None:
        return ""
    return str(value).strip()


# ---------------------------------------------------------------------------
# CSV parsing
# ---------------------------------------------------------------------------

def _parse_csv_bytes(file_content: bytes) -> tuple[list[list], str]:
    """
    Parse CSV bytes with encoding fallback.
    Returns (list_of_rows, encoding_used).
    """
    import csv

    encodings_to_try = ["utf-8", "utf-8-sig", "windows-1252", "latin-1"]

    for encoding in encodings_to_try:
        try:
            text = file_content.decode(encoding)
            reader = csv.reader(io.StringIO(text))
            rows = [row for row in reader]
            return rows, encoding
        except (UnicodeDecodeError, Exception):
            continue

    raise ParseError(
        "CSV file could not be decoded with any supported encoding",
        "parse_failed",
    )


# ---------------------------------------------------------------------------
# xlsx parsing
# ---------------------------------------------------------------------------

def _parse_xlsx_bytes(file_content: bytes) -> tuple[list[list[list]], list[str]]:
    """
    Parse xlsx bytes.
    Returns (list_of_sheets_as_row_lists, sheet_names).
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_content),
            data_only=True,
        )
    except Exception as e:
        raise ParseError(f"Could not parse xlsx file: {e}", "parse_failed")

    sheet_names = wb.sheetnames
    sheets = []
    for ws in wb.worksheets:
        sheet_rows = []
        for row in ws.iter_rows(values_only=True):
            sheet_rows.append(list(row))
        sheets.append(sheet_rows)

    return sheets, sheet_names


# ---------------------------------------------------------------------------
# xls parsing
# ---------------------------------------------------------------------------

def _parse_xls_bytes(file_content: bytes) -> tuple[list[list[list]], list[str]]:
    """
    Parse xls bytes using xlrd.
    Returns (list_of_sheets_as_row_lists, sheet_names).
    """
    try:
        import xlrd
    except ImportError:
        raise ParseError("xlrd is not installed; cannot parse .xls files", "parse_failed")

    try:
        wb = xlrd.open_workbook(file_contents=file_content)
    except Exception as e:
        raise ParseError(f"Could not parse xls file: {e}", "parse_failed")

    sheet_names = wb.sheet_names()
    sheets = []
    for ws in wb.sheets():
        sheet_rows = []
        for row_idx in range(ws.nrows):
            row = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                # xlrd cell types: 0=empty, 1=text, 2=number, 3=date, 4=bool, 5=error
                import xlrd as _xlrd
                if cell.ctype == _xlrd.XL_CELL_EMPTY:
                    row.append(None)
                elif cell.ctype == _xlrd.XL_CELL_NUMBER:
                    # Return int if whole number, else float
                    v = cell.value
                    row.append(int(v) if v == int(v) else v)
                elif cell.ctype == _xlrd.XL_CELL_DATE:
                    row.append(str(cell.value))
                else:
                    row.append(cell.value)
        sheet_rows.append(row)
        sheets.append(sheet_rows)

    return sheets, sheet_names


# ---------------------------------------------------------------------------
# Metadata period extraction
# ---------------------------------------------------------------------------

# Label sets for extracting reporting period start/end from metadata rows.
_PERIOD_START_LABELS: frozenset[str] = frozenset({
    "reporting period start",
    "period start",
    "from",
    "start date",
    "period from",
})

_PERIOD_END_LABELS: frozenset[str] = frozenset({
    "reporting period end",
    "period end",
    "through",
    "end date",
    "period through",
    "to",
    "period to",
})


def _extract_metadata_periods(
    raw_rows: list[list],
    header_idx: int,
) -> tuple[Optional[str], Optional[str]]:
    """
    Scan rows before the detected header (indices 0..header_idx-1) for period
    label/value pairs and return (period_start, period_end).

    For each row, each cell (lowercased and stripped) is checked against the
    known start and end label sets.  When a match is found the next cell in
    the same row is used as the value.

    Returns:
        A 2-tuple (start_value, end_value).  Either or both may be None if
        the corresponding label was not found.
    """
    period_start: Optional[str] = None
    period_end: Optional[str] = None

    for row in raw_rows[:header_idx]:
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            label = str(cell).strip().lower()
            # Look at the next cell in the same row for the value
            next_idx = col_idx + 1
            if next_idx >= len(row):
                continue
            next_cell = row[next_idx]
            if next_cell is None or str(next_cell).strip() == "":
                continue
            value = str(next_cell).strip()

            if label in _PERIOD_START_LABELS and period_start is None:
                period_start = value
            elif label in _PERIOD_END_LABELS and period_end is None:
                period_end = value

    return period_start, period_end


# ---------------------------------------------------------------------------
# Core parse_upload
# ---------------------------------------------------------------------------

def parse_upload(file_content: bytes, filename: str) -> ParsedSheet:
    """
    Parse an uploaded spreadsheet file and return structured data.

    Args:
        file_content: Raw bytes of the uploaded file.
        filename: Original filename (used to determine file type).

    Returns:
        ParsedSheet with detected columns, sample rows, and data row count.

    Raises:
        ParseError: If the file type is unsupported or the file cannot be parsed.
    """
    ext = _get_extension(filename)

    if ext not in SUPPORTED_EXTENSIONS:
        raise ParseError(
            f"Unsupported file type '{ext}'. Upload a .xlsx, .xls, or .csv file.",
            "unsupported_file_type",
        )

    # Parse to raw rows
    if ext == ".xlsx":
        sheets, sheet_names = _parse_xlsx_bytes(file_content)
        # Use first sheet; fall back to second if first has < 3 rows
        active_sheet_idx = 0
        if len(sheets) > 1 and len(sheets[0]) < 3:
            active_sheet_idx = 1
        raw_rows = sheets[active_sheet_idx]
        sheet_name = sheet_names[active_sheet_idx]

    elif ext == ".xls":
        sheets, sheet_names = _parse_xls_bytes(file_content)
        active_sheet_idx = 0
        if len(sheets) > 1 and len(sheets[0]) < 3:
            active_sheet_idx = 1
        raw_rows = sheets[active_sheet_idx]
        sheet_name = sheet_names[active_sheet_idx]

    else:  # .csv
        raw_rows, _ = _parse_csv_bytes(file_content)
        # Convert list-of-lists of strings (csv reader returns strings)
        raw_rows = [row for row in raw_rows]
        sheet_name = "Sheet1"

    if not raw_rows:
        raise ParseError("File is empty", "parse_failed")

    # Detect header row
    header_idx = _detect_header_row(raw_rows)
    header_row = raw_rows[header_idx]

    # Extract metadata periods from rows before the detected header
    metadata_period_start, metadata_period_end = _extract_metadata_periods(raw_rows, header_idx)

    # Forward-fill None headers (merged header cells)
    filled_headers: list[Optional[str]] = []
    last_header = None
    for cell in header_row:
        if cell is None or str(cell).strip() == "":
            filled_headers.append(last_header)
        else:
            last_header = str(cell).strip()
            filled_headers.append(last_header)

    # Build column names (deduplicate if needed)
    column_names: list[str] = []
    seen_names: dict[str, int] = {}
    for h in filled_headers:
        name = h if h else f"Column_{len(column_names) + 1}"
        if name in seen_names:
            seen_names[name] += 1
            name = f"{name}_{seen_names[name]}"
        else:
            seen_names[name] = 0
        column_names.append(name)

    n_cols = len(column_names)

    # Get raw data rows (after header)
    raw_data = raw_rows[header_idx + 1:]

    # Mark which rows were originally empty BEFORE any normalization/forward-fill
    originally_empty = [_row_is_all_empty(row) for row in raw_data]

    # Normalize to n_cols length
    data_raw = [
        (list(row) + [None] * n_cols)[:n_cols]
        for row in raw_data
    ]
    # Forward-fill each column to handle merged cells (openpyxl merged cell regions)
    for col_idx in range(n_cols):
        data_raw = _forward_fill_column(data_raw, col_idx)

    # Filter out empty rows and summary rows
    data_rows_list: list[dict] = []
    found_summary = False
    for idx, row in enumerate(data_raw):
        # Skip rows that were originally empty (before forward-fill)
        if originally_empty[idx]:
            continue
        if found_summary:
            # Everything after first summary row is non-data
            continue
        if _is_summary_row(row):
            found_summary = True
            continue
        # Build dict
        row_dict = {column_names[i]: _cell_to_str(row[i]) for i in range(n_cols)}
        data_rows_list.append(row_dict)

    total_rows = len(raw_rows) - 1  # subtract header
    sample = data_rows_list[:5]

    return ParsedSheet(
        column_names=column_names,
        all_rows=data_rows_list,
        sample_rows=sample,
        data_rows=len(data_rows_list),
        sheet_name=sheet_name,
        total_rows=total_rows,
        metadata_period_start=metadata_period_start,
        metadata_period_end=metadata_period_end,
    )


# ---------------------------------------------------------------------------
# apply_mapping
# ---------------------------------------------------------------------------

def apply_mapping(
    parsed: ParsedSheet,
    column_mapping: dict[str, str],
) -> MappedData:
    """
    Apply a confirmed column mapping to parsed sheet data and aggregate results.

    Derives net_sales via:
      1. Direct mapping if a column is mapped to "net_sales".
      2. Derived: gross_sales - returns (if both are mapped and net_sales is not).
      3. Gross as net: if only gross_sales is mapped and returns is not.

    Args:
        parsed: Result from parse_upload().
        column_mapping: Maps detected column names to canonical Likha field names.

    Returns:
        MappedData with aggregated net_sales, category_sales, and other fields.

    Raises:
        MappingError: If net_sales cannot be determined or the result is negative.
    """
    # Reverse map: field_name -> list of column names mapped to it
    # "metadata" and "ignore" are both excluded from calculation field grouping.
    field_to_columns: dict[str, list[str]] = {}
    metadata_cols: list[str] = []
    for col, field in column_mapping.items():
        if field == "metadata":
            metadata_cols.append(col)
        elif field and field != "ignore":
            field_to_columns.setdefault(field, []).append(col)

    has_net_sales_col = "net_sales" in field_to_columns
    has_gross_sales_col = "gross_sales" in field_to_columns
    has_returns_col = "returns" in field_to_columns
    has_category_col = "product_category" in field_to_columns

    # Determine how we will compute net_sales
    if not has_net_sales_col and not has_gross_sales_col:
        raise MappingError(
            "No column is mapped to 'Net Sales'. Map a column to 'Net Sales' before confirming.",
            "net_sales_column_required",
        )

    # Aggregate data
    net_sales_total = Decimal("0")
    gross_sales_total = Decimal("0")
    returns_total = Decimal("0")
    category_sales: dict[str, Decimal] = {}
    licensee_royalty_total = Decimal("0")
    has_royalty_values = False
    # Metadata: col_name -> list of raw string values collected across all rows
    metadata_values: dict[str, list[str]] = {col: [] for col in metadata_cols}

    net_sales_cols = field_to_columns.get("net_sales", [])
    gross_sales_cols = field_to_columns.get("gross_sales", [])
    returns_cols = field_to_columns.get("returns", [])
    category_cols = field_to_columns.get("product_category", [])
    royalty_cols = field_to_columns.get("licensee_reported_royalty", [])

    for row in parsed.all_rows:
        # Category for this row
        row_category: Optional[str] = None
        if category_cols:
            cat_val = row.get(category_cols[0], "").strip()
            if cat_val:
                row_category = cat_val

        # Net sales for this row
        row_net = Decimal("0")
        if has_net_sales_col:
            for col in net_sales_cols:
                val = _to_decimal_safe(row.get(col))
                if val is not None:
                    row_net += val
        elif has_gross_sales_col:
            row_gross = Decimal("0")
            for col in gross_sales_cols:
                val = _to_decimal_safe(row.get(col))
                if val is not None:
                    row_gross += val
            row_ret = Decimal("0")
            if has_returns_col:
                for col in returns_cols:
                    val = _to_decimal_safe(row.get(col))
                    if val is not None:
                        row_ret += val
            row_net = row_gross - row_ret
            gross_sales_total += row_gross
            returns_total += row_ret

        net_sales_total += row_net

        if has_category_col and row_category:
            category_sales[row_category] = category_sales.get(row_category, Decimal("0")) + row_net

        # Gross/returns aggregation (for direct net_sales path)
        if has_net_sales_col and has_gross_sales_col:
            for col in gross_sales_cols:
                val = _to_decimal_safe(row.get(col))
                if val is not None:
                    gross_sales_total += val
        if has_net_sales_col and has_returns_col:
            for col in returns_cols:
                val = _to_decimal_safe(row.get(col))
                if val is not None:
                    returns_total += val

        # Licensee reported royalty
        for col in royalty_cols:
            val = _to_decimal_safe(row.get(col))
            if val is not None:
                licensee_royalty_total += val
                has_royalty_values = True

        # Metadata: collect raw cell values (pass-through, no calculation)
        for col in metadata_cols:
            cell_val = _cell_to_str(row.get(col))
            metadata_values[col].append(cell_val)

    # Validate net_sales
    if net_sales_total < Decimal("0"):
        raise MappingError(
            f"Net sales aggregated to a negative value (${net_sales_total}). "
            "Verify the returns column is mapped correctly.",
            "negative_net_sales",
        )

    return MappedData(
        net_sales=net_sales_total,
        category_sales=category_sales if category_sales else None,
        licensee_reported_royalty=licensee_royalty_total if has_royalty_values else None,
        gross_sales=gross_sales_total if gross_sales_total else None,
        returns=returns_total if returns_total else None,
        metadata=metadata_values if metadata_cols else None,
    )


# ---------------------------------------------------------------------------
# extract_cross_check_values
# ---------------------------------------------------------------------------

def extract_cross_check_values(
    parsed: ParsedSheet,
    column_mapping: dict[str, str],
) -> dict[str, Optional[str]]:
    """
    Extract the first non-null value from each cross-check column in the mapping.

    Cross-check fields are informational only (licensee_name, report_period,
    royalty_rate).  They are not used in calculation but are compared against
    contract data during the confirm step to produce upload warnings.

    Args:
        parsed: Result from parse_upload().
        column_mapping: Maps detected column names to canonical Likha field names.

    Returns:
        Dict with keys "licensee_name", "report_period", "royalty_rate".
        Each value is the first non-empty string found in the mapped column,
        or None if the field is not mapped or all values are empty.
    """
    # Build reverse map: field_name -> first column mapped to it
    field_to_col: dict[str, str] = {}
    for col, field in column_mapping.items():
        if field in CROSS_CHECK_FIELDS and field not in field_to_col:
            field_to_col[field] = col

    result: dict[str, Optional[str]] = {
        "licensee_name": None,
        "report_period": None,
        "royalty_rate": None,
    }

    for field_name, col in field_to_col.items():
        for row in parsed.all_rows:
            val = row.get(col, "")
            if val and str(val).strip():
                result[field_name] = str(val).strip()
                break

    return result


# ---------------------------------------------------------------------------
# claude_suggest
# ---------------------------------------------------------------------------

# Timeout in seconds for the Claude AI column mapping call.
_AI_MAPPING_TIMEOUT = 5


def claude_suggest(
    columns: list[dict],
    contract_context: dict,
) -> dict[str, str]:
    """
    Ask Claude to suggest canonical field names for unresolved spreadsheet columns.

    This is a best-effort function: it returns an empty dict on any error
    (timeout, API failure, bad JSON, missing API key).  Callers must never
    depend on it succeeding.

    Args:
        columns: List of {"name": str, "samples": list[str]} dicts for each
                 unresolved column.
        contract_context: Dict with keys licensee_name, royalty_base,
                          has_categories, and categories.

    Returns:
        Dict mapping column name -> canonical field name for columns Claude
        could classify.  Invalid field names are silently discarded.
        Returns {} on any failure.
    """
    if not columns:
        return {}

    try:
        import anthropic
        import httpx

        api_key = os.getenv("ANTHROPIC_API_KEY")

        valid_fields_list = sorted(VALID_FIELDS)

        prompt = (
            "You are a royalty-report data analyst.\n"
            "Given the contract context and a list of spreadsheet columns "
            "(each with sample values), map every column to exactly one "
            "canonical field name from the valid_fields list.\n\n"
            f"Contract context:\n{json.dumps(contract_context, indent=2)}\n\n"
            f"Valid field names: {json.dumps(valid_fields_list)}\n\n"
            f"Columns to classify:\n{json.dumps(columns, indent=2)}\n\n"
            "Respond with ONLY a JSON object mapping column name to field name. "
            "Example: {\"Rev\": \"net_sales\", \"Sku Group\": \"product_category\"}. "
            "Do not include any explanation or markdown."
        )

        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=512,
            timeout=_AI_MAPPING_TIMEOUT,
            messages=[{"role": "user", "content": prompt}],
        )

        raw_text = response.content[0].text.strip()

        # Strip markdown code fences if present
        if raw_text.startswith("```"):
            lines = raw_text.split("\n")
            lines = [ln for ln in lines if not ln.strip().startswith("```")]
            raw_text = "\n".join(lines).strip()

        parsed: dict = json.loads(raw_text)

        # Discard any field values that are not in VALID_FIELDS
        return {
            col: field_val
            for col, field_val in parsed.items()
            if field_val in VALID_FIELDS
        }

    except Exception:
        logger.debug("claude_suggest: silent fallback", exc_info=True)
        return {}


# ---------------------------------------------------------------------------
# suggest_mapping
# ---------------------------------------------------------------------------

def suggest_mapping(
    column_names: list[str],
    saved_mapping: Optional[dict[str, str]],
    contract_context: Optional[dict] = None,
    return_source: bool = False,
    sample_rows: Optional[list[dict]] = None,
) -> "dict[str, str] | tuple[dict[str, str], str, dict[str, str]]":
    """
    Suggest a column mapping based on keyword synonyms and/or a saved mapping.

    When a saved_mapping is provided, columns present in the saved mapping
    use their saved value.  New columns (not in the saved mapping) fall back
    to keyword matching.

    When contract_context is provided, columns that keyword matching could
    not resolve (mapped to "ignore") are sent to Claude AI for a second-pass
    suggestion.  If Claude is unavailable or times out the result falls back
    gracefully to the keyword-only output.

    Args:
        column_names: List of detected column names from the uploaded file.
        saved_mapping: Optional previously-saved mapping for this licensee.
        contract_context: Optional dict with contract metadata for AI mapping.
                          Keys: licensee_name, royalty_base, has_categories,
                          categories.  When None, AI step is skipped.
        return_source: When True, return a 3-tuple (mapping, source_str,
                       col_sources) instead of just the mapping dict.
                       source_str is one of: "saved", "ai", "suggested",
                       "none".  col_sources maps every column name to its
                       individual source: "saved", "keyword", "ai", "none".
        sample_rows: Optional list of sample data rows (dicts keyed by column
                     name) used to extract per-column sample values that are
                     sent to Claude.  When None or empty, samples default to [].

    Returns:
        When return_source is False (default): dict mapping each column name
        to a canonical Likha field name.
        When return_source is True: (mapping_dict, source_str, col_sources)
        3-tuple.
    """
    result: dict[str, str] = {}
    col_sources: dict[str, str] = {}
    any_saved = False

    for col in column_names:
        # 1. Check saved mapping first
        if saved_mapping and col in saved_mapping:
            result[col] = saved_mapping[col]
            col_sources[col] = "saved"
            any_saved = True
            continue

        # 2. Keyword synonym matching (case-insensitive, substring)
        normalized = col.lower().strip()
        # Prepend a space to support the ' ns' synonym check against leading space
        padded = " " + normalized

        matched_field = "ignore"
        for field_name, synonyms in FIELD_SYNONYMS.items():
            for synonym in synonyms:
                syn_lower = synonym.lower()
                # Check both the normalized name and the padded name
                if syn_lower in normalized or syn_lower in padded:
                    matched_field = field_name
                    break
            if matched_field != "ignore":
                break

        result[col] = matched_field
        col_sources[col] = "keyword" if matched_field != "ignore" else "none"

    # 3. AI second-pass: only when contract_context is provided
    ai_resolved_any = False
    if contract_context is not None:
        unresolved = [
            {
                "name": col,
                "samples": _extract_column_samples(col, sample_rows),
            }
            for col, val in result.items()
            if val == "ignore" and (not saved_mapping or col not in saved_mapping)
        ]

        if unresolved:
            ai_suggestions = claude_suggest(unresolved, contract_context)
            for col, field_val in ai_suggestions.items():
                if col in result and result[col] == "ignore":
                    result[col] = field_val
                    col_sources[col] = "ai"
                    ai_resolved_any = True

    if not return_source:
        return result

    # Determine overall source label
    all_ignore = all(v == "ignore" for v in result.values())
    if any_saved:
        source = "saved"
    elif ai_resolved_any:
        source = "ai"
    elif all_ignore:
        source = "none"
    else:
        source = "suggested"

    return result, source, col_sources


def _extract_column_samples(col: str, sample_rows: Optional[list[dict]]) -> list[str]:
    """Return up to 5 non-empty string values for *col* from *sample_rows*."""
    if not sample_rows:
        return []
    samples: list[str] = []
    for row in sample_rows:
        val = row.get(col)
        if val is None:
            continue
        s = str(val).strip()
        if s:
            samples.append(s)
        if len(samples) == 5:
            break
    return samples


# ---------------------------------------------------------------------------
# claude_suggest_categories
# ---------------------------------------------------------------------------

def claude_suggest_categories(
    report_categories: list[str],
    contract_categories: list[str],
) -> dict[str, str]:
    """
    Ask Claude to map report category names to contract category names.

    This is a best-effort function: returns {} on any failure (timeout, API
    failure, bad JSON, missing API key). Callers must never depend on it
    succeeding.

    Args:
        report_categories: Category names found in the uploaded report.
        contract_categories: Canonical category names from the contract.

    Returns:
        Dict mapping report_category -> contract_category.
        Only entries where the suggested contract_category is in
        contract_categories are included. Returns {} on any failure.
    """
    if not report_categories:
        return {}

    try:
        import anthropic

        api_key = os.getenv("ANTHROPIC_API_KEY")

        prompt = (
            "You are a royalty-report data analyst.\n"
            "A licensee's sales report uses different category names than the contract.\n"
            "Map each report category to the single best-matching contract category.\n\n"
            f"Contract categories: {json.dumps(contract_categories)}\n\n"
            f"Report categories to map: {json.dumps(report_categories)}\n\n"
            "Respond with ONLY a JSON object mapping each report category to a contract "
            "category. Use exactly the contract category name as it appears above. "
            "Example: {\"Tops & Bottoms\": \"Apparel\", \"Footwear\": \"Footwear\"}. "
            "Do not include any explanation or markdown."
        )

        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=512,
            timeout=_AI_MAPPING_TIMEOUT,
            messages=[{"role": "user", "content": prompt}],
        )

        raw_text = response.content[0].text.strip()

        # Strip markdown code fences if present
        if raw_text.startswith("```"):
            lines = raw_text.split("\n")
            lines = [ln for ln in lines if not ln.strip().startswith("```")]
            raw_text = "\n".join(lines).strip()

        parsed_response: dict = json.loads(raw_text)

        # Only keep entries where the suggested category is a valid contract category
        valid_set = set(contract_categories)
        return {
            report_cat: contract_cat
            for report_cat, contract_cat in parsed_response.items()
            if contract_cat in valid_set
        }

    except Exception:
        logger.debug("claude_suggest_categories: silent fallback", exc_info=True)
        return {}


# ---------------------------------------------------------------------------
# suggest_category_mapping
# ---------------------------------------------------------------------------

def suggest_category_mapping(
    report_categories: list[str],
    contract_categories: list[str],
    saved_category_mapping: Optional[dict[str, str]],
) -> tuple[dict[str, str], dict[str, str]]:
    """
    Suggest a mapping from report category names to contract category names.

    Resolution order for each report category:
    1. Saved alias (from licensee_column_mappings.category_mapping)
    2. Exact match (case-insensitive)
    3. Substring match (report cat contains contract cat or vice versa)
    4. AI suggestion (Claude, for remaining unresolved)

    Args:
        report_categories: Distinct category values found in the uploaded file.
        contract_categories: Canonical category names from the contract's royalty_rate.
        saved_category_mapping: Previously saved aliases for this licensee, or None.

    Returns:
        A 2-tuple of:
          - mapping: dict mapping report_category -> contract_category (only
            resolved entries; unresolved categories are absent or None).
          - sources: dict mapping report_category -> source string, one of:
            "saved", "exact", "substring", "ai", "none".
    """
    result: dict[str, str] = {}
    sources: dict[str, str] = {}
    unresolved: list[str] = []

    contract_lower: dict[str, str] = {c.lower(): c for c in contract_categories}

    for report_cat in report_categories:
        # 1. Saved alias
        if saved_category_mapping and report_cat in saved_category_mapping:
            result[report_cat] = saved_category_mapping[report_cat]
            sources[report_cat] = "saved"
            continue

        normalized = report_cat.lower().strip()

        # 2. Exact match (case-insensitive)
        if normalized in contract_lower:
            result[report_cat] = contract_lower[normalized]
            sources[report_cat] = "exact"
            continue

        # 3. Substring match
        matched: Optional[str] = None
        for contract_cat_lower, contract_cat in contract_lower.items():
            if contract_cat_lower in normalized or normalized in contract_cat_lower:
                matched = contract_cat
                break

        if matched is not None:
            result[report_cat] = matched
            sources[report_cat] = "substring"
            continue

        # 4. Needs AI
        unresolved.append(report_cat)
        sources[report_cat] = "none"

    # AI pass for unresolved categories
    if unresolved:
        ai_suggestions = claude_suggest_categories(unresolved, contract_categories)
        for report_cat, contract_cat in ai_suggestions.items():
            if report_cat in sources and sources[report_cat] == "none":
                result[report_cat] = contract_cat
                sources[report_cat] = "ai"

    return result, sources
